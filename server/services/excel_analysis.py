from __future__ import annotations

import math
import re
import unicodedata
import uuid
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


FIELD_ALIASES = {
    "date": ["日期", "销售日期", "交易日期", "订单日期", "date", "sales date", "order date"],
    "region": ["地区", "区域", "销售区域", "市场", "region", "area", "market"],
    "product": ["产品", "商品", "产品名称", "商品名称", "product", "product name", "item"],
    "sales_amount": [
        "销售额", "销售金额", "金额", "营业收入", "收入", "sales", "revenue", "amount", "sales amount",
    ],
    "quantity": ["数量", "销量", "销售数量", "quantity", "qty", "units"],
}
REQUIRED_FIELD = "sales_amount"
HEADER_SCAN_ROWS = 12


class ExcelAnalysisError(Exception):
    def __init__(self, code: str, message: str, status_code: int = 400, **context: object):
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code
        self.context = context

    def as_detail(self) -> dict[str, object]:
        return {"code": self.code, "message": self.message, **self.context}


@dataclass(frozen=True)
class SheetCandidate:
    name: str
    header_row: int
    rows: int
    columns: int
    density: float
    matched_fields: tuple[str, ...]
    header_values: tuple[str, ...]
    score: float

    def as_dict(self) -> dict[str, object]:
        return {
            "name": self.name,
            "headerRow": self.header_row + 1,
            "rows": self.rows,
            "columns": self.columns,
            "nonEmptyRows": self.rows,
            "density": round(self.density, 4),
            "matchedFields": list(self.matched_fields),
            "score": round(self.score, 3),
        }


def normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[\s_\-—–/\\:：()（）\[\]【】]+", "", text)


NORMALIZED_ALIASES = {
    field: {normalize_header(alias) for alias in aliases}
    for field, aliases in FIELD_ALIASES.items()
}


def _header_matches(values: list[object]) -> dict[str, list[str]]:
    matches: dict[str, list[str]] = {field: [] for field in FIELD_ALIASES}
    for value in values:
        original = str(value).strip() if value is not None else ""
        normalized = normalize_header(value)
        if not normalized:
            continue
        for field, aliases in NORMALIZED_ALIASES.items():
            if normalized in aliases:
                matches[field].append(original)
    return {field: columns for field, columns in matches.items() if columns}


def _sheet_candidates(path: Path) -> list[SheetCandidate]:
    try:
        sheets = pd.read_excel(path, sheet_name=None, header=None, engine="openpyxl")
    except Exception as error:
        raise ExcelAnalysisError("INVALID_XLSX", "无法读取 XLSX 工作簿。") from error

    candidates: list[SheetCandidate] = []
    for name, raw in sheets.items():
        raw = raw.dropna(axis=1, how="all")
        if raw.empty or raw.dropna(how="all").empty:
            candidates.append(SheetCandidate(
                name=str(name), header_row=0, rows=0, columns=0, density=0,
                matched_fields=(), header_values=(), score=0,
            ))
            continue
        best_header = 0
        best_matches: dict[str, list[str]] = {}
        for index in range(min(HEADER_SCAN_ROWS, len(raw))):
            matches = _header_matches(raw.iloc[index].tolist())
            if len(matches) > len(best_matches):
                best_header = index
                best_matches = matches
        data = raw.iloc[best_header + 1 :].dropna(how="all")
        columns = max(1, int(raw.shape[1]))
        rows = int(len(data))
        density = 0.0 if rows == 0 else float(data.notna().sum().sum() / (rows * columns))
        required_bonus = 25 if REQUIRED_FIELD in best_matches else 0
        score = len(best_matches) * 20 + required_bonus + min(rows, 1000) / 100 + density * 10
        candidates.append(SheetCandidate(
            name=str(name),
            header_row=best_header,
            rows=rows,
            columns=columns,
            density=density,
            matched_fields=tuple(best_matches),
            header_values=tuple(str(value).strip() if value is not None else "" for value in raw.iloc[best_header]),
            score=score,
        ))
    return candidates


def _select_sheet(candidates: list[SheetCandidate], requested: str | None) -> SheetCandidate:
    valid_candidates = [candidate for candidate in candidates if candidate.rows > 0]
    if not valid_candidates:
        raise ExcelAnalysisError("EMPTY_WORKBOOK", "工作簿中没有可分析的数据。")
    if requested is not None:
        selected = next((candidate for candidate in valid_candidates if candidate.name == requested), None)
        if selected is None:
            raise ExcelAnalysisError(
                "NO_VALID_SHEET",
                "指定的工作表不存在或没有数据。",
                availableSheets=[candidate.name for candidate in valid_candidates],
            )
        return selected
    if len(valid_candidates) == 1:
        return valid_candidates[0]

    ranked = sorted(valid_candidates, key=lambda candidate: candidate.score, reverse=True)
    top = ranked[0]
    runner_up = ranked[1]
    confident = (
        REQUIRED_FIELD in top.matched_fields
        and len(top.matched_fields) >= 2
        and (top.score - runner_up.score >= 12 or REQUIRED_FIELD not in runner_up.matched_fields)
    )
    if not confident:
        raise ExcelAnalysisError(
            "NEEDS_SHEET_SELECTION",
            "检测到多个可能的数据工作表，请选择要分析的 sheet。",
            409,
            needsSheetSelection=True,
            candidates=[candidate.as_dict() for candidate in ranked],
        )
    return top


def _field_mapping(columns: list[object]) -> dict[str, dict[str, object]]:
    mapping: dict[str, dict[str, object]] = {}
    ambiguous: dict[str, list[str]] = {}
    for field, aliases in NORMALIZED_ALIASES.items():
        matched = [str(column) for column in columns if normalize_header(column) in aliases]
        if len(matched) > 1:
            ambiguous[field] = matched
        elif matched:
            mapping[field] = {"column": matched[0], "confidence": 1.0, "match": "alias_exact"}
    if ambiguous:
        raise ExcelAnalysisError(
            "AMBIGUOUS_FIELD_MAPPING",
            "部分字段匹配到多个候选列，请明确列名后重试。",
            ambiguousFields=ambiguous,
            availableColumns=[str(column) for column in columns],
        )
    if REQUIRED_FIELD not in mapping:
        raise ExcelAnalysisError(
            "MISSING_REQUIRED_FIELD",
            "未找到销售额字段，无法计算销售指标。",
            missingFields=[REQUIRED_FIELD],
            availableColumns=[str(column) for column in columns],
        )
    return mapping


def parse_number(value: object) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = re.sub(r"[¥￥$€£,，\s]", "", text)
    if not re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)", text):
        return None
    number = float(text)
    return -number if negative else number


def _formula_warnings(path: Path, sheet_name: str) -> list[str]:
    formula_book = load_workbook(path, read_only=True, data_only=False)
    value_book = load_workbook(path, read_only=True, data_only=True)
    missing_cache = 0
    try:
        for formula_row, value_row in zip(
            formula_book[sheet_name].iter_rows(), value_book[sheet_name].iter_rows(),
        ):
            for formula_cell, value_cell in zip(formula_row, value_row):
                if formula_cell.data_type == "f" and value_cell.value is None:
                    missing_cache += 1
    finally:
        formula_book.close()
        value_book.close()
    return (
        [f"检测到 {missing_cache} 个公式单元格没有缓存结果；openpyxl 不会执行公式，这些值按缺失处理。"]
        if missing_cache else []
    )


def _ranked_groups(frame: pd.DataFrame, group_column: str, sales_column: str) -> list[dict[str, object]]:
    grouped = frame.dropna(subset=[group_column, sales_column]).copy()
    grouped[group_column] = grouped[group_column].astype(str).str.strip()
    grouped = grouped[grouped[group_column] != ""]
    if grouped.empty:
        return []
    totals = grouped.groupby(group_column, dropna=True)[sales_column].sum().sort_values(ascending=False)
    denominator = float(totals.sum())
    return [
        {
            "name": str(name),
            "sales": float(amount),
            "share": (float(amount) / denominator if denominator else None),
            "rank": rank,
        }
        for rank, (name, amount) in enumerate(totals.items(), start=1)
    ]


def analyze_excel(path: Path, metadata: dict[str, object], requested_sheet: str | None) -> dict[str, object]:
    candidates = _sheet_candidates(path)
    selected = _select_sheet(candidates, requested_sheet)
    try:
        raw = pd.read_excel(path, sheet_name=selected.name, header=selected.header_row, engine="openpyxl")
    except Exception as error:
        raise ExcelAnalysisError("ANALYSIS_FAILED", "读取选定工作表失败。", 500) from error

    raw.columns = [str(column).strip() for column in raw.columns]
    header_matches = _header_matches(list(selected.header_values))
    duplicate_header_matches = {
        field: columns for field, columns in header_matches.items() if len(columns) > 1
    }
    if duplicate_header_matches:
        raise ExcelAnalysisError(
            "AMBIGUOUS_FIELD_MAPPING",
            "部分字段匹配到多个候选列，请明确列名后重试。",
            ambiguousFields=duplicate_header_matches,
            availableColumns=list(selected.header_values),
        )
    empty_rows_removed = int(raw.isna().all(axis=1).sum())
    non_empty = raw.dropna(how="all").copy()
    raw_row_count = int(len(raw))
    duplicate_rows = int(non_empty.duplicated(keep="first").sum())
    clean = non_empty.drop_duplicates(keep="first").copy()
    mapping = _field_mapping(list(clean.columns))
    missing_by_column = {str(column): int(clean[column].isna().sum()) for column in clean.columns}
    missing_cells = int(sum(missing_by_column.values()))

    sales_column = str(mapping["sales_amount"]["column"])
    clean["__sales"] = clean[sales_column].map(parse_number)
    invalid_sales_rows = int(clean["__sales"].isna().sum())
    valid_sales = clean.dropna(subset=["__sales"]).copy()
    if valid_sales.empty:
        raise ExcelAnalysisError("NO_VALID_SALES_ROWS", "销售额列中没有可用于计算的有效数据。")

    date_column = str(mapping["date"]["column"]) if "date" in mapping else None
    invalid_date_rows = 0
    if date_column:
        clean["__date"] = pd.to_datetime(clean[date_column], errors="coerce", format="mixed")
        invalid_date_rows = int(clean[date_column].notna().sum() - clean["__date"].notna().sum())
        valid_sales["__date"] = pd.to_datetime(
            valid_sales[date_column], errors="coerce", format="mixed",
        )

    sales_series = valid_sales["__sales"].astype(float)
    monthly: list[dict[str, object]] = []
    date_range = None
    yoy_growth = None
    warnings = _formula_warnings(path, selected.name)
    if date_column:
        valid_dates = valid_sales.dropna(subset=["__date"]).copy()
        if not valid_dates.empty:
            valid_dates["__month"] = valid_dates["__date"].dt.to_period("M").astype(str)
            monthly_groups = valid_dates.groupby("__month")["__sales"].agg(["sum", "count"])
            monthly = [
                {"month": str(month), "sales": float(row["sum"]), "validRowCount": int(row["count"])}
                for month, row in monthly_groups.iterrows()
            ]
            minimum = valid_dates["__date"].min()
            maximum = valid_dates["__date"].max()
            date_range = {"start": minimum.date().isoformat(), "end": maximum.date().isoformat()}
            years = sorted(valid_dates["__date"].dt.year.unique())
            if len(years) >= 2:
                previous, latest = years[-2], years[-1]
                year_month = valid_dates.assign(
                    __year=valid_dates["__date"].dt.year,
                    __month_number=valid_dates["__date"].dt.month,
                )
                comparable_months = (
                    set(year_month.loc[year_month["__year"] == previous, "__month_number"])
                    & set(year_month.loc[year_month["__year"] == latest, "__month_number"])
                )
                if comparable_months:
                    previous_total = float(year_month.loc[
                        (year_month["__year"] == previous) & year_month["__month_number"].isin(comparable_months),
                        "__sales",
                    ].sum())
                    latest_total = float(year_month.loc[
                        (year_month["__year"] == latest) & year_month["__month_number"].isin(comparable_months),
                        "__sales",
                    ].sum())
                    if previous_total != 0:
                        yoy_growth = (latest_total - previous_total) / abs(previous_total)
                    else:
                        warnings.append("YoY unavailable: previous comparable period total is zero")
                else:
                    warnings.append("YoY unavailable: insufficient comparable periods")
            else:
                warnings.append("YoY unavailable: insufficient comparable periods")
        else:
            warnings.append("日期列没有可用于趋势分析的有效值。")
    else:
        warnings.append("未识别到日期列，月度趋势和同比不可用。")

    region_column = str(mapping["region"]["column"]) if "region" in mapping else None
    product_column = str(mapping["product"]["column"]) if "product" in mapping else None
    regional_sales = _ranked_groups(valid_sales, region_column, "__sales") if region_column else []
    product_sales = _ranked_groups(valid_sales, product_column, "__sales") if product_column else []
    if not region_column:
        warnings.append("未识别到地区列，地区分析不可用。")
    if not product_column:
        warnings.append("未识别到产品列，产品分析不可用。")

    quantity = None
    if "quantity" in mapping:
        quantity_column = str(mapping["quantity"]["column"])
        quantity_values = clean[quantity_column].map(parse_number).dropna()
        quantity = {"totalQuantity": float(quantity_values.sum()), "validQuantityRowCount": int(len(quantity_values))}

    return {
        "analysisId": f"analysis_{uuid.uuid4().hex}",
        "fileId": metadata["fileId"],
        "fileName": metadata["fileName"],
        "workbook": {
            "sheetCount": len(candidates),
            "sheets": [candidate.as_dict() for candidate in candidates],
            "selectedSheet": selected.name,
            "selectionMethod": "requested" if requested_sheet else (
                "single_valid_sheet" if len([candidate for candidate in candidates if candidate.rows > 0]) == 1 else "scored"
            ),
        },
        "dataset": {
            "sheetName": selected.name,
            "headerRow": selected.header_row + 1,
            "rawRowCount": raw_row_count,
            "cleanRowCount": int(len(clean)),
            "columnCount": int(len(raw.columns)),
            "columns": [str(column) for column in raw.columns],
            "dateRange": date_range,
        },
        "fieldMapping": mapping,
        "dataQuality": {
            "rawRowCount": raw_row_count,
            "cleanRowCount": int(len(clean)),
            "emptyRowsRemoved": empty_rows_removed,
            "duplicateRows": duplicate_rows,
            "duplicateRowsRemoved": duplicate_rows,
            "missingCells": missing_cells,
            "missingByColumn": missing_by_column,
            "invalidSalesRows": invalid_sales_rows,
            "invalidDateRows": invalid_date_rows,
            "warnings": list(warnings),
        },
        "metrics": {
            "sales": {
                "totalSales": float(sales_series.sum()),
                "averageSales": float(sales_series.mean()),
                "medianSales": float(sales_series.median()),
                "minSales": float(sales_series.min()),
                "maxSales": float(sales_series.max()),
                "validSalesRowCount": int(len(sales_series)),
                "yoyGrowth": float(yoy_growth) if yoy_growth is not None else None,
            },
            "topRegion": regional_sales[0] if regional_sales else None,
            "topProduct": product_sales[0] if product_sales else None,
            "quantity": quantity,
        },
        "monthlyTrend": monthly,
        "regionalSales": regional_sales,
        "productSales": product_sales,
        "warnings": warnings,
        "summary": "已完成真实 Excel 的结构化解析与指标计算；DeepSeek 智能洞察在独立区域生成和展示。",
    }
