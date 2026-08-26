"""Generate a deterministic, synthetic sales workbook for TaskPet A1 analysis."""

from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = ["销售日期", "区域", "产品名称", "销售额", "销量", "订单编号", "渠道"]
REGIONS = ["华北", "华东", "华南", "西南"]
PRODUCTS = ["示例产品 A", "示例产品 B", "示例产品 C"]
CHANNELS = ["线上", "线下", "合作渠道"]


def synthetic_rows() -> list[list[object]]:
    rows: list[list[object]] = []
    for month in range(1, 7):
        for index, region in enumerate(REGIONS):
            product = PRODUCTS[(month + index) % len(PRODUCTS)]
            channel = CHANNELS[(month * 2 + index) % len(CHANNELS)]
            quantity = 8 + month * 3 + index * 2
            unit_price = 120 + ((month + index) % 4) * 35
            rows.append([
                date(2026, month, 5 + index * 6),
                region,
                product,
                quantity * unit_price,
                quantity,
                f"DEMO-2026{month:02d}-{index + 1:03d}",
                channel,
            ])
    return rows


def build_workbook(output: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthetic Sales"
    sheet.append(HEADERS)
    for row in synthetic_rows():
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in sheet["D"][1:]:
        cell.number_format = "#,##0.00"
    for cell in sheet["E"][1:]:
        cell.number_format = "#,##0"

    widths = {"A": 14, "B": 12, "C": 18, "D": 14, "E": 10, "F": 20, "G": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

    table = Table(displayName="SyntheticSalesTable", ref=f"A1:G{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("taskpet_demo_sales.xlsx"),
        help="Output .xlsx path (default: taskpet_demo_sales.xlsx)",
    )
    args = parser.parse_args()
    output = build_workbook(args.output.resolve())
    print(f"Generated synthetic workbook: {output}")


if __name__ == "__main__":
    main()
