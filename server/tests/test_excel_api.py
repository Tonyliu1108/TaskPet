import io
import sys
from datetime import datetime
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import main
from services.excel_analysis import ExcelAnalysisError, analyze_excel, parse_number
from services.upload_storage import UploadSettings


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def workbook_bytes(rows, headers=None, sheet_name="销售数据"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_workbook(tmp_path, payload, name="source.xlsx"):
    path = tmp_path / name
    path.write_bytes(payload)
    return path


def metadata(file_id="file_00000000000000000000000000000000"):
    return {"fileId": file_id, "fileName": "2026年销售数据.xlsx"}


@pytest.mark.asyncio
async def test_upload_valid_xlsx_returns_persistent_file_id(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 25 * 1024 * 1024))
    payload = workbook_bytes([["2026-01-01", "华南", "A", 100]], ["日期", "地区", "产品", "销售额"])
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.post(
            "/api/files/upload",
            files={"file": ("2026年销售数据.xlsx", payload, XLSX_MIME)},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["fileId"].startswith("file_")
    assert body["fileName"] == "2026年销售数据.xlsx"
    assert body["size"] == len(payload)
    assert (tmp_path / f'{body["fileId"]}.xlsx').is_file()
    assert (tmp_path / f'{body["fileId"]}.json').is_file()


@pytest.mark.asyncio
async def test_uploaded_file_metadata_survives_api_state_restart(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 25 * 1024 * 1024))
    payload = workbook_bytes([["2026-01-01", 100]], ["日期", "销售额"])
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        upload = await client.post(
            "/api/files/upload",
            files={"file": ("持久化测试.xlsx", payload, XLSX_MIME)},
        )
        main.analysis_cache.clear()
        metadata_response = await client.get(f'/api/files/{upload.json()["fileId"]}')
        analysis_response = await client.post(
            "/api/excel/analyze",
            json={"fileId": upload.json()["fileId"], "sheetName": None},
        )

    assert upload.status_code == 200
    assert metadata_response.status_code == 200
    assert metadata_response.json() == {**upload.json(), "contentType": XLSX_MIME, "status": "available"}
    assert analysis_response.status_code == 200
    assert analysis_response.json()["metrics"]["sales"]["totalSales"] == 100


@pytest.mark.asyncio
async def test_uploaded_file_metadata_returns_structured_not_found(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 25 * 1024 * 1024))
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.get("/api/files/file_00000000000000000000000000000000")

    assert response.status_code == 404
    assert response.json()["detail"]["code"] == "FILE_NOT_FOUND"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("name", "payload", "content_type", "expected_code"),
    [
        ("sales.csv", b"a,b\n1,2", XLSX_MIME, "INVALID_FILE_TYPE"),
        ("sales.xlsx", b"not-a-zip", XLSX_MIME, "INVALID_XLSX"),
        ("sales.xlsx", b"not-an-xlsx", "text/plain", "INVALID_FILE_TYPE"),
    ],
)
async def test_upload_rejects_invalid_files(
    tmp_path, monkeypatch, name, payload, content_type, expected_code,
):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 25 * 1024 * 1024))
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.post(
            "/api/files/upload", files={"file": (name, payload, content_type)},
        )
    assert response.status_code == 400
    assert response.json()["detail"]["code"] == expected_code
    assert list(tmp_path.iterdir()) == []


@pytest.mark.asyncio
async def test_upload_rejects_oversized_file(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 32))
    payload = workbook_bytes([[1]], ["销售额"])
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.post(
            "/api/files/upload",
            files={"file": ("sales.xlsx", payload, XLSX_MIME)},
        )
    assert response.status_code == 413
    assert response.json()["detail"]["code"] == "FILE_TOO_LARGE"


def test_parse_number_supports_currency_spaces_and_parentheses():
    assert parse_number("1,234.56") == 1234.56
    assert parse_number("¥1,234.56") == 1234.56
    assert parse_number("￥1,234") == 1234
    assert parse_number("1 234") == 1234
    assert parse_number("(1,234.00)") == -1234
    assert parse_number("not money") is None


def test_analysis_computes_exact_metrics_and_quality(tmp_path):
    headers = ["销售日期", "销售区域", "产品名称", "销售金额", "数量", "备注"]
    rows = [
        ["2026-01-01", "华南", "A", "¥1,000.00", 2, None],
        ["2026/01/15", "华北", "B", "2,000", 3, "ok"],
        ["2026-02-01", "华南", "A", "(500)", 1, None],
        ["bad-date", "华南", "B", "300", 4, None],
        ["2026-02-02", "华东", "C", "invalid", 5, None],
        [None, None, None, None, None, None],
        ["2026-01-01", "华南", "A", "¥1,000.00", 2, None],
    ]
    path = save_workbook(tmp_path, workbook_bytes(rows, headers))
    result = analyze_excel(path, metadata(), None)

    assert result["workbook"]["selectedSheet"] == "销售数据"
    assert result["fieldMapping"]["sales_amount"]["column"] == "销售金额"
    assert result["dataQuality"] == {
        "rawRowCount": 7,
        "cleanRowCount": 5,
        "emptyRowsRemoved": 1,
        "duplicateRows": 1,
        "duplicateRowsRemoved": 1,
        "missingCells": 4,
        "missingByColumn": {"销售日期": 0, "销售区域": 0, "产品名称": 0, "销售金额": 0, "数量": 0, "备注": 4},
        "invalidSalesRows": 1,
        "invalidDateRows": 1,
        "warnings": ["YoY unavailable: insufficient comparable periods"],
    }
    sales = result["metrics"]["sales"]
    assert sales["totalSales"] == 2800
    assert sales["averageSales"] == 700
    assert sales["medianSales"] == 650
    assert sales["minSales"] == -500
    assert sales["maxSales"] == 2000
    assert sales["validSalesRowCount"] == 4
    assert sales["yoyGrowth"] is None
    assert result["monthlyTrend"] == [
        {"month": "2026-01", "sales": 3000.0, "validRowCount": 2},
        {"month": "2026-02", "sales": -500.0, "validRowCount": 1},
    ]
    assert result["regionalSales"][0] == {"name": "华北", "sales": 2000.0, "share": 2000 / 2800, "rank": 1}
    assert result["productSales"][0] == {"name": "B", "sales": 2300.0, "share": 2300 / 2800, "rank": 1}
    assert result["metrics"]["quantity"] == {"totalQuantity": 15.0, "validQuantityRowCount": 5}


def test_english_headers_and_optional_dimensions(tmp_path):
    path = save_workbook(tmp_path, workbook_bytes(
        [[datetime(2025, 1, 1), "North", "Widget", "$10"], ["2026-01-01", "North", "Widget", "$15"]],
        ["Order Date", "Region", "Product Name", "Revenue"],
    ))
    result = analyze_excel(path, metadata(), None)
    assert result["metrics"]["sales"]["totalSales"] == 25
    assert result["metrics"]["sales"]["yoyGrowth"] == 0.5
    assert set(result["fieldMapping"]) == {"date", "region", "product", "sales_amount"}


def test_multiple_sheets_requires_selection_when_scores_are_close(tmp_path):
    workbook = Workbook()
    first = workbook.active
    first.title = "华南"
    first.append(["日期", "地区", "产品", "销售额"])
    first.append(["2026-01-01", "华南", "A", 100])
    second = workbook.create_sheet("华北")
    second.append(["日期", "地区", "产品", "销售额"])
    second.append(["2026-01-01", "华北", "A", 200])
    output = io.BytesIO()
    workbook.save(output)
    path = save_workbook(tmp_path, output.getvalue())

    with pytest.raises(ExcelAnalysisError) as error_info:
        analyze_excel(path, metadata(), None)
    assert error_info.value.code == "NEEDS_SHEET_SELECTION"
    assert len(error_info.value.context["candidates"]) == 2

    result = analyze_excel(path, metadata(), "华北")
    assert result["metrics"]["sales"]["totalSales"] == 200


def test_multiple_sheets_auto_selects_clear_sales_sheet_and_reports_all_sheets(tmp_path):
    workbook = Workbook()
    notes = workbook.active
    notes.title = "说明"
    notes.append(["这是工作簿说明"])
    notes.append(["数据口径以销售明细为准"])
    sales = workbook.create_sheet("Sales Data")
    sales.append(["Order Date", "Region", "Product", "Sales Amount"])
    sales.append(["2026-01-01", "North", "A", 100])
    empty = workbook.create_sheet("空白")
    output = io.BytesIO()
    workbook.save(output)
    path = save_workbook(tmp_path, output.getvalue(), "scored.xlsx")

    result = analyze_excel(path, metadata(), None)
    assert result["workbook"]["selectedSheet"] == "Sales Data"
    assert result["workbook"]["selectionMethod"] == "scored"
    assert result["workbook"]["sheetCount"] == 3
    assert [sheet["name"] for sheet in result["workbook"]["sheets"]] == ["说明", "Sales Data", "空白"]
    assert result["metrics"]["sales"]["totalSales"] == 100


def test_empty_workbook_and_missing_sales_field_are_structured_errors(tmp_path):
    empty = Workbook()
    empty_path = tmp_path / "empty.xlsx"
    empty.save(empty_path)
    with pytest.raises(ExcelAnalysisError) as empty_error:
        analyze_excel(empty_path, metadata(), None)
    assert empty_error.value.code == "EMPTY_WORKBOOK"

    missing_path = save_workbook(
        tmp_path,
        workbook_bytes([["2026-01-01", "华南", "A"]], ["日期", "地区", "产品"]),
        "missing.xlsx",
    )
    with pytest.raises(ExcelAnalysisError) as missing_error:
        analyze_excel(missing_path, metadata(), None)
    assert missing_error.value.code == "MISSING_REQUIRED_FIELD"
    assert missing_error.value.context["missingFields"] == ["sales_amount"]


def test_duplicate_sales_headers_are_not_silently_guessed(tmp_path):
    path = save_workbook(
        tmp_path,
        workbook_bytes([[100, 200]], ["销售额", "金额"]),
        "ambiguous.xlsx",
    )
    with pytest.raises(ExcelAnalysisError) as error_info:
        analyze_excel(path, metadata(), None)
    assert error_info.value.code == "AMBIGUOUS_FIELD_MAPPING"
    assert error_info.value.context["ambiguousFields"]["sales_amount"] == ["销售额", "金额"]


def test_no_valid_sales_rows_and_formula_cache_warning(tmp_path):
    invalid_path = save_workbook(
        tmp_path,
        workbook_bytes([["2026-01-01", "bad"]], ["日期", "销售额"]),
        "invalid-sales.xlsx",
    )
    with pytest.raises(ExcelAnalysisError) as invalid_error:
        analyze_excel(invalid_path, metadata(), None)
    assert invalid_error.value.code == "NO_VALID_SALES_ROWS"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售数据"
    sheet.append(["日期", "销售额"])
    sheet.append(["2026-01-01", 100])
    sheet.append(["2026-01-02", "=50+50"])
    formula_path = tmp_path / "formula.xlsx"
    workbook.save(formula_path)
    result = analyze_excel(formula_path, metadata(), None)
    assert result["metrics"]["sales"]["totalSales"] == 100
    assert result["dataQuality"]["invalidSalesRows"] == 1
    assert any("公式单元格没有缓存结果" in warning for warning in result["warnings"])


def test_requested_missing_sheet_returns_no_valid_sheet(tmp_path):
    path = save_workbook(tmp_path, workbook_bytes([[100]], ["销售额"]))
    with pytest.raises(ExcelAnalysisError) as error_info:
        analyze_excel(path, metadata(), "不存在")
    assert error_info.value.code == "NO_VALID_SHEET"


@pytest.mark.asyncio
async def test_analyze_api_uses_file_id_and_returns_file_not_found(monkeypatch, tmp_path):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 25 * 1024 * 1024))
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.post(
            "/api/excel/analyze",
            json={"fileId": "file_00000000000000000000000000000000", "sheetName": None},
        )
    assert response.status_code == 404
    assert response.json()["detail"]["code"] == "FILE_NOT_FOUND"


@pytest.mark.asyncio
async def test_upload_then_analyze_api_returns_exact_real_bytes_result(monkeypatch, tmp_path):
    monkeypatch.setattr(main, "upload_settings", UploadSettings(tmp_path, 25 * 1024 * 1024))
    payload = workbook_bytes(
        [["2026-03-01", "华南", "A", 120], ["2026-03-02", "华北", "B", 80]],
        ["日期", "地区", "产品", "销售额"],
    )
    transport = httpx.ASGITransport(app=main.app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        upload = await client.post(
            "/api/files/upload",
            files={"file": ("真实测试.xlsx", payload, XLSX_MIME)},
        )
        result = await client.post(
            "/api/excel/analyze",
            json={"fileId": upload.json()["fileId"], "sheetName": None},
        )
    assert upload.status_code == 200
    assert result.status_code == 200
    assert result.json()["metrics"]["sales"]["totalSales"] == 200
    assert result.json()["monthlyTrend"] == [
        {"month": "2026-03", "sales": 200.0, "validRowCount": 2},
    ]
